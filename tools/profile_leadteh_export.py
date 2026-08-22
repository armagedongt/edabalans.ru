from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def normalized(value: object) -> str:
    return text(value).casefold()


def duplicate_summary(values: list[str]) -> tuple[int, int, int]:
    counts = Counter(value for value in values if value)
    return len(values), len(counts), sum(1 for count in counts.values() if count > 1)


def split_tags(raw: str) -> list[str]:
    raw = raw.strip().strip("[]")
    if not raw:
        return []
    return [item.strip(" \t\r\n\"'") for item in re.split(r"[,;\n]+", raw) if item.strip(" \t\r\n\"'")]


def profile(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    sheet.reset_dimensions()
    rows = sheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    header_groups: dict[str, list[int]] = {}
    for index, header in enumerate(headers):
        header_groups.setdefault(header.casefold(), []).append(index)

    lead_ids: list[str] = []
    emails: list[str] = []
    telegram_ids: list[str] = []
    usernames: list[str] = []
    phones: list[str] = []
    tag_sets = 0
    tags: Counter[str] = Counter()
    messenger: Counter[str] = Counter()
    stages: Counter[str] = Counter()
    nonempty: Counter[int] = Counter()
    row_count = 0

    def values_for(row: tuple[object, ...], *names: str) -> list[str]:
        result: list[str] = []
        for name in names:
            for index in header_groups.get(name.casefold(), []):
                if index < len(row):
                    value = text(row[index])
                    if value:
                        result.append(value)
        return result

    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        row_count += 1
        for index, value in enumerate(row):
            if value not in (None, ""):
                nonempty[index] += 1
        lead_ids.extend(values_for(row, "ID"))
        emails.extend(normalized(value) for value in values_for(row, "Email"))
        telegram_ids.extend(normalized(value) for value in values_for(row, "Telegram ID", "TG_ID"))
        usernames.extend(normalized(value).lstrip("@") for value in values_for(row, "Username"))
        phones.extend(re.sub(r"\D", "", value) for value in values_for(row, "Телефон"))
        messenger.update(values_for(row, "Мессенджер"))
        stages.update(values_for(row, "Этап"))
        raw_tags = values_for(row, "Теги")
        if raw_tags:
            tag_sets += 1
            for raw in raw_tags:
                tags.update(split_tags(raw))

    print(f"FILE {path.name}")
    print(f"rows={row_count}")
    for label, values in (
        ("lead_id", lead_ids),
        ("email_values", emails),
        ("telegram_id_values", telegram_ids),
        ("username_values", usernames),
        ("phone_values", phones),
    ):
        total, unique, duplicate_keys = duplicate_summary(values)
        print(f"{label}: values={total}, unique={unique}, duplicated_keys={duplicate_keys}")
    print(f"rows_with_tags={tag_sets}, unique_tag_spellings={len(tags)}, tag_assignments={sum(tags.values())}")
    print("messengers=" + repr(messenger.most_common()))
    if stages:
        print("stages=" + repr(stages.most_common()))
    print("top_tags=")
    for name, count in tags.most_common(100):
        print(f"  {count}\t{name}")
    print("populated_columns=")
    for index, count in sorted(nonempty.items(), key=lambda item: (-item[1], item[0])):
        if count:
            print(f"  {count}\t{index + 1}\t{headers[index] or '<empty>'}")
    workbook.close()


if __name__ == "__main__":
    for argument in sys.argv[1:]:
        profile(Path(argument))
