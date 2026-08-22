from __future__ import annotations

import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import psycopg
from openpyxl import load_workbook


def text(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value or "").strip().split())


def tag_key(value: str) -> str:
    value = unicodedata.normalize("NFKC", value).casefold().replace("ё", "е")
    value = re.sub(r"[^0-9a-zа-я]+", " ", value)
    return " ".join(value.split())


def split_tags(raw: str) -> list[str]:
    raw = raw.strip().strip("[]")
    return [item.strip(" \t\r\n\"'") for item in re.split(r"[,;\n]+", raw) if item.strip(" \t\r\n\"'")]


def category(name: str) -> str:
    value = tag_key(name)
    if any(word in value for word in ("подпис", " пдп", "непдп", "не пдп")):
        return "Подписка"
    if any(word in value for word in ("рассыл", "воронк", "дожим", "цепоч", "этап", "партия")):
        return "Рассылки и воронки"
    if value.startswith("пост ") or any(word in value for word in ("смотрел", "видел", "открыл", "нажал", "день ")):
        return "Контент и действия"
    if any(word in value for word in ("купил", "оплат", "тариф", "цена", "покуп")):
        return "Покупки"
    if value.startswith("из ") or any(word in value for word in ("источник", "пикабу", "директ")):
        return "Источники"
    if "лотере" in value:
        return "Лотерея"
    if value in {"пусто", "старый", "первое посещение"} or re.fullmatch(r"[a-z0-9]{10,}", value):
        return "Служебное/мусор"
    return "Прочее"


def main(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    sheet.reset_dimensions()
    rows = sheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    positions: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(headers):
        positions[header.casefold()].append(index)

    def values(row: tuple[object, ...], *names: str) -> list[str]:
        found: list[str] = []
        for name in names:
            for index in positions.get(name.casefold(), []):
                if index < len(row):
                    value = text(row[index])
                    if value:
                        found.append(value)
        return found

    database_url = os.environ["DATABASE_URL"].replace("postgresql+psycopg://", "postgresql://", 1)
    with psycopg.connect(database_url) as connection:
        telegram_by_id: dict[str, set[str]] = defaultdict(set)
        users_by_username: dict[str, set[str]] = defaultdict(set)
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT user_id::text, platform_user_id, username FROM messenger_accounts WHERE platform = 'telegram'"
            )
            for user_id, platform_id, username in cursor:
                if platform_id:
                    telegram_by_id[text(platform_id)].add(user_id)
                if username:
                    users_by_username[text(username).casefold().lstrip("@")].add(user_id)
            cursor.execute("SELECT user_id::text, email_normalized FROM user_emails")
            users_by_email = {email.casefold(): user_id for user_id, email in cursor if email}
            cursor.execute("SELECT DISTINCT user_id::text FROM payments WHERE payment_status = 'paid' AND user_id IS NOT NULL")
            buyer_ids = {row[0] for row in cursor}

    export_rows = 0
    exact_rows = 0
    exact_users: set[str] = set()
    matched_buyers: set[str] = set()
    conflicts = 0
    username_only_candidates = 0
    no_identity = 0
    tag_counts: Counter[str] = Counter()

    for row in rows:
        if not any(value not in (None, "") for value in row):
            continue
        export_rows += 1
        messenger = (values(row, "Мессенджер") or [""])[0].casefold()
        telegram_ids = values(row, "Telegram ID", "TG_ID") if messenger == "telegram" else []
        emails = [email.casefold() for email in values(row, "Email") if "@" in email]
        usernames = [name.casefold().lstrip("@") for name in values(row, "Username")]

        telegram_matches = set().union(*(telegram_by_id.get(item, set()) for item in telegram_ids)) if telegram_ids else set()
        email_matches = {users_by_email[item] for item in emails if item in users_by_email}
        exact = telegram_matches | email_matches
        if telegram_matches and email_matches and telegram_matches != email_matches:
            conflicts += 1
        if exact:
            exact_rows += 1
            exact_users.update(exact)
            matched_buyers.update(exact & buyer_ids)
        elif any(len(users_by_username.get(item, set())) == 1 for item in usernames):
            username_only_candidates += 1
        elif not telegram_ids and not emails and not usernames:
            no_identity += 1

        for raw in values(row, "Теги"):
            tag_counts.update(split_tags(raw))

    normalized_groups: dict[str, list[str]] = defaultdict(list)
    for name in tag_counts:
        normalized_groups[tag_key(name)].append(name)
    duplicate_groups = [sorted(names, key=str.casefold) for names in normalized_groups.values() if len(names) > 1]

    category_names: dict[str, set[str]] = defaultdict(set)
    category_assignments: Counter[str] = Counter()
    for name, count in tag_counts.items():
        group = category(name)
        category_names[group].add(name)
        category_assignments[group] += count

    print(f"export_rows={export_rows}")
    print(f"exactly_matched_rows={exact_rows}")
    print(f"exactly_matched_existing_users={len(exact_users)}")
    print(f"matched_existing_buyers={len(matched_buyers)}")
    print(f"new_or_unmatched_rows={export_rows - exact_rows}")
    print(f"identity_conflicts={conflicts}")
    print(f"username_only_candidates={username_only_candidates}")
    print(f"rows_without_tg_email_username={no_identity}")
    print(f"normalized_duplicate_tag_groups={len(duplicate_groups)}")
    for names in sorted(duplicate_groups, key=lambda item: tag_key(item[0])):
        print("DUPLICATE " + " | ".join(f"{name} ({tag_counts[name]})" for name in names))
    print("TAG_CATEGORIES")
    for name in sorted(category_names):
        print(f"{name}: names={len(category_names[name])}, assignments={category_assignments[name]}")
    workbook.close()


if __name__ == "__main__":
    main(Path(sys.argv[1]))
