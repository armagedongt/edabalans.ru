from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook


def clean_header(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def audit(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    print(f"FILE {path.name}")
    for sheet in workbook.worksheets:
        # LeadTeh exports an incorrect A1 worksheet dimension even when the
        # underlying XML contains many rows. Reset it before streaming cells.
        sheet.reset_dimensions()
        rows = sheet.iter_rows(values_only=True)
        headers = [clean_header(value) for value in next(rows, ())]
        populated_rows = sum(1 for row in rows if any(value not in (None, "") for value in row))
        print(f"SHEET {sheet.title!r}: rows={populated_rows}, columns={len(headers)}")
        for index, header in enumerate(headers, start=1):
            print(f"  {index}: {header or '<empty>'}")
    workbook.close()


if __name__ == "__main__":
    for argument in sys.argv[1:]:
        audit(Path(argument))
