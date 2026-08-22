from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import unicodedata
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import psycopg
from openpyxl import load_workbook
from psycopg.types.json import Jsonb


SOURCE = "leadteh_contacts_20260822"


def clean(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value or "").strip().split())


def normalize_email(value: str) -> str:
    return value.strip().casefold()


def normalize_phone(value: str) -> str:
    digits = re.sub(r"\D", "", value)
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    return digits


def parse_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value
    raw = clean(value)
    if not raw:
        return None
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        return None


def split_tags(raw: str) -> list[str]:
    raw = raw.strip().strip("[]")
    return [item.strip(" \t\r\n\"'") for item in re.split(r"[,;\n]+", raw) if item.strip(" \t\r\n\"'")]


def tag_key(value: str) -> str:
    value = unicodedata.normalize("NFKC", value).casefold().replace("ё", "е")
    value = re.sub(r"[^0-9a-zа-я]+", " ", value)
    return " ".join(value.split())


def tag_category(name: str) -> str:
    value = tag_key(name)
    if any(word in value for word in ("подпис", " пдп", "непдп", "не пдп")):
        return "subscription"
    if any(word in value for word in ("рассыл", "воронк", "дожим", "цепоч", "этап", "партия")):
        return "mailing_funnel"
    if value.startswith("пост ") or any(word in value for word in ("смотрел", "видел", "открыл", "нажал", "день ")):
        return "content_action"
    if any(word in value for word in ("купил", "оплат", "тариф", "цена", "покуп")):
        return "purchase_signal"
    if value.startswith("из ") or any(word in value for word in ("источник", "пикабу", "директ")):
        return "source"
    if "лотере" in value:
        return "lottery"
    if value in {"пусто", "старый", "первое посещение"} or re.fullmatch(r"[a-z0-9]{10,}", value):
        return "technical"
    return "other"


def tag_code(name: str) -> str:
    return "leadteh_" + hashlib.sha256(name.encode("utf-8")).hexdigest()[:32]


def database_url() -> str:
    return os.environ["DATABASE_URL"].replace("postgresql+psycopg://", "postgresql://", 1)


def import_contacts(path: Path, source: str, file_sha256: str, dry_run: bool = False) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    sheet.reset_dimensions()
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    positions: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(headers):
        positions[header.casefold()].append(index)

    def values(row: tuple[object, ...], *names: str) -> list[str]:
        found: list[str] = []
        for name in names:
            for index in positions.get(name.casefold(), []):
                if index < len(row):
                    value = clean(row[index])
                    if value:
                        found.append(value)
        return found

    counters = defaultdict(int)
    assignments: list[tuple[uuid.UUID, uuid.UUID]] = []
    batch_id = uuid.uuid4()

    with psycopg.connect(database_url()) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                "INSERT INTO import_batches (id, source, status) VALUES (%s, %s, 'running')",
                (batch_id, source),
            )
            cursor.execute(
                "SELECT row_hash FROM legacy_import_records WHERE source = %s", (source,)
            )
            imported_hashes = {row[0] for row in cursor}

            telegram_by_id: dict[str, uuid.UUID] = {}
            cursor.execute(
                "SELECT platform_user_id, user_id FROM messenger_accounts "
                "WHERE platform = 'telegram' AND platform_user_id IS NOT NULL"
            )
            for platform_id, user_id in cursor:
                telegram_by_id[clean(platform_id)] = user_id

            email_by_value: dict[str, uuid.UUID] = {}
            cursor.execute("SELECT email_normalized, user_id FROM user_emails")
            for email, user_id in cursor:
                email_by_value[email.casefold()] = user_id

            tag_by_name: dict[str, uuid.UUID] = {}
            cursor.execute("SELECT name, id FROM tags")
            for name, tag_id in cursor:
                tag_by_name[name] = tag_id

            for row_number, row in enumerate(rows, start=2):
                if not any(value not in (None, "") for value in row):
                    continue
                counters["rows"] += 1
                lead_id = (values(row, "ID") or [str(row_number)])[0]
                row_hash = hashlib.sha256(lead_id.encode("utf-8")).hexdigest()
                if row_hash in imported_hashes:
                    counters["duplicates"] += 1
                    continue

                name = (values(row, "Имя") or [""])[0]
                messenger = (values(row, "Мессенджер") or [""])[0].casefold()
                username = (values(row, "Username") or [""])[0].lstrip("@")
                telegram_id = ""
                if messenger == "telegram":
                    telegram_id = (values(row, "Telegram ID", "TG_ID") or [""])[0]
                emails = [normalize_email(item) for item in values(row, "Email") if "@" in item]
                phone_original = (values(row, "Телефон") or [""])[0]
                phone_normalized = normalize_phone(phone_original)
                created_at = parse_datetime((values(row, "Дата создания") or [""])[0])
                source_raw = (values(row, "Источник:", "Источник: Без источника") or [""])[0]
                first_activity = parse_datetime((values(row, "Первая активность") or [""])[0])
                raw_tags: list[str] = []
                for raw in values(row, "Теги"):
                    raw_tags.extend(split_tags(raw))

                matched_ids: set[uuid.UUID] = set()
                if telegram_id and telegram_id in telegram_by_id:
                    matched_ids.add(telegram_by_id[telegram_id])
                matched_ids.update(email_by_value[email] for email in emails if email in email_by_value)

                if len(matched_ids) > 1:
                    counters["needs_review"] += 1
                    cursor.execute(
                        """
                        INSERT INTO legacy_import_records
                            (import_batch_id, source, source_row_number, row_hash,
                             external_record_id, status, reason, raw_payload)
                        VALUES (%s, %s, %s, %s, %s, 'needs_review', %s, %s)
                        """,
                        (
                            batch_id,
                            source,
                            row_number,
                            row_hash,
                            lead_id,
                            "Telegram ID и email указывают на разных пользователей",
                            Jsonb(
                                {
                                    "leadteh_id": lead_id,
                                    "messenger": messenger,
                                    "telegram_id": telegram_id,
                                    "emails": emails,
                                    "username": username,
                                }
                            ),
                        ),
                    )
                    imported_hashes.add(row_hash)
                    continue

                if matched_ids:
                    user_id = next(iter(matched_ids))
                    counters["matched"] += 1
                    reason = "matched_existing"
                    cursor.execute(
                        "UPDATE users SET display_name = COALESCE(display_name, %s), "
                        "first_seen_at = COALESCE(first_seen_at, %s), updated_at = now() WHERE id = %s",
                        (name or None, created_at, user_id),
                    )
                else:
                    user_id = uuid.uuid4()
                    cursor.execute(
                        """
                        INSERT INTO users
                            (id, display_name, status, data_origin, first_seen_at)
                        VALUES (%s, %s, 'active', 'legacy_import', %s)
                        """,
                        (user_id, name or None, created_at),
                    )
                    counters["created"] += 1
                    reason = "created_new"

                for email in dict.fromkeys(emails):
                    owner = email_by_value.get(email)
                    if owner is None:
                        cursor.execute(
                            """
                            INSERT INTO user_emails
                                (user_id, email_original, email_normalized,
                                 verification_status, source, first_seen_at)
                            VALUES (%s, %s, %s, 'legacy_unverified', 'leadteh_legacy', %s)
                            """,
                            (user_id, email, email, created_at),
                        )
                        email_by_value[email] = user_id
                        counters["emails"] += 1

                if messenger:
                    platform_id = telegram_id or None
                    if platform_id:
                        owner = telegram_by_id.get(platform_id)
                        if owner is None:
                            cursor.execute(
                                """
                                INSERT INTO messenger_accounts
                                    (user_id, platform, platform_user_id, username,
                                     first_name, source, first_seen_at, last_seen_at)
                                VALUES (%s, %s, %s, %s, %s, 'leadteh_legacy', %s, %s)
                                """,
                                (user_id, messenger, platform_id, username or None, name or None, created_at, created_at),
                            )
                            telegram_by_id[platform_id] = user_id
                            counters["messengers"] += 1
                        else:
                            cursor.execute(
                                "UPDATE messenger_accounts SET username = COALESCE(username, %s), "
                                "first_name = COALESCE(first_name, %s) "
                                "WHERE platform = %s AND platform_user_id = %s",
                                (username or None, name or None, messenger, platform_id),
                            )
                    elif username:
                        cursor.execute(
                            """
                            INSERT INTO messenger_accounts
                                (user_id, platform, platform_user_id, username,
                                 first_name, source, first_seen_at, last_seen_at)
                            SELECT %s, %s, NULL, %s, %s, 'leadteh_legacy', %s, %s
                            WHERE NOT EXISTS (
                                SELECT 1 FROM messenger_accounts
                                WHERE user_id = %s AND platform = %s AND username = %s
                            )
                            """,
                            (user_id, messenger, username, name or None, created_at, created_at,
                             user_id, messenger, username),
                        )

                if phone_normalized:
                    cursor.execute(
                        """
                        INSERT INTO user_phones
                            (user_id, phone_original, phone_normalized, source)
                        VALUES (%s, %s, %s, 'leadteh_legacy')
                        ON CONFLICT (user_id, phone_normalized) DO NOTHING
                        """,
                        (user_id, phone_original, phone_normalized),
                    )
                    counters["phones"] += cursor.rowcount

                if source_raw or first_activity:
                    cursor.execute(
                        """
                        INSERT INTO attribution_events
                            (user_id, import_batch_id, event_type, source_raw, occurred_at)
                        VALUES (%s, %s, 'leadteh_legacy', %s, %s)
                        """,
                        (user_id, batch_id, source_raw or None, first_activity or created_at),
                    )
                    counters["attribution"] += 1

                for tag_name in dict.fromkeys(raw_tags):
                    tag_id = tag_by_name.get(tag_name)
                    if tag_id is None:
                        tag_id = uuid.uuid4()
                        cursor.execute(
                            """
                            INSERT INTO tags (id, code, name, category, status)
                            VALUES (%s, %s, %s, %s, 'active')
                            """,
                            (tag_id, tag_code(tag_name), tag_name, tag_category(tag_name)),
                        )
                        tag_by_name[tag_name] = tag_id
                        counters["tags_created"] += 1
                    assignments.append((user_id, tag_id))

                cursor.execute(
                    """
                    INSERT INTO legacy_import_records
                        (import_batch_id, source, source_row_number, row_hash,
                         external_record_id, status, user_id, reason, raw_payload)
                    VALUES (%s, %s, %s, %s, %s, 'imported', %s, %s, %s)
                    """,
                    (
                        batch_id,
                        source,
                        row_number,
                        row_hash,
                        lead_id,
                        user_id,
                        reason,
                        Jsonb(
                            {
                                "leadteh_id": lead_id,
                                "messenger": messenger,
                                "tag_count": len(set(raw_tags)),
                                "export_sha256": file_sha256,
                            }
                        ),
                    ),
                )
                imported_hashes.add(row_hash)

            cursor.execute(
                "CREATE TEMP TABLE stage_leadteh_user_tags (user_id uuid, tag_id uuid) ON COMMIT DROP"
            )
            with cursor.copy("COPY stage_leadteh_user_tags (user_id, tag_id) FROM STDIN") as copy:
                for user_id, tag_id in assignments:
                    copy.write_row((user_id, tag_id))
            cursor.execute(
                """
                INSERT INTO user_tags (id, user_id, tag_id, source)
                SELECT gen_random_uuid(), user_id, tag_id, 'leadteh_legacy'
                FROM stage_leadteh_user_tags
                ON CONFLICT (user_id, tag_id) DO NOTHING
                """
            )
            counters["tag_assignments"] = cursor.rowcount
            counters["distinct_tags"] = len(tag_by_name)
            summary = dict(counters)
            cursor.execute(
                "UPDATE import_batches SET status = %s, finished_at = now(), summary = %s WHERE id = %s",
                ("dry_run" if dry_run else "completed", Jsonb(summary), batch_id),
            )
            if dry_run:
                connection.rollback()
            else:
                connection.commit()
    workbook.close()
    return dict(counters)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import LeadTeh contacts into CRM")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--source", default=SOURCE)
    parser.add_argument("--sha256", required=True)
    parser.add_argument("--dry-run", action="store_true")
    arguments = parser.parse_args()
    print(json.dumps(import_contacts(arguments.xlsx, arguments.source, arguments.sha256, arguments.dry_run), ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
